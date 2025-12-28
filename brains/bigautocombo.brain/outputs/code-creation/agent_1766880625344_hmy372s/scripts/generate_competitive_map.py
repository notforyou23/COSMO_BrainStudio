#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, re
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SEED = ROOT / "data" / "competitive_landscape" / "competitors_seed.csv"
DEFAULT_OUT_DIR = ROOT / "outputs"
DEFAULT_XLSX = DEFAULT_OUT_DIR / "competitive_map.xlsx"
DEFAULT_CSV = DEFAULT_OUT_DIR / "competitive_map.csv"
DEFAULT_MD = DEFAULT_OUT_DIR / "competitive_summary.md"

SEED_FALLBACK = [
  {"competitor":"ExampleKit Co","supplier_type":"Kit Supplier","region":"NA","pricing_band":"Mid","price_range_usd":"$2,000-$5,000","warranty_term":"12 months","compliance_posture":"CE; RoHS","differentiators":"Fast ship kits; modular","notes":"Replace with seed CSV"},
  {"competitor":"IntegratorPro","supplier_type":"Integrator","region":"NA/EU","pricing_band":"High","price_range_usd":"$10,000-$50,000","warranty_term":"12 months","compliance_posture":"ISO 9001; project documentation","differentiators":"Turnkey installs; training","notes":""},
  {"competitor":"OEM Program X","supplier_type":"OEM Program","region":"Global","pricing_band":"High","price_range_usd":"Varies","warranty_term":"24 months","compliance_posture":"UL; CE; RoHS","differentiators":"Brand channel; enterprise support","notes":""},
]

def _s(x): return (x or "").strip()
def _norm_type(t):
  t=_s(t).lower()
  if "kit" in t: return "Kit Supplier"
  if "integr" in t: return "Integrator"
  if "oem" in t: return "OEM Program"
  if t in {"supplier","vendor"}: return "Kit Supplier"
  return _s(t).title() or "Unknown"
def _norm_pricing(p):
  p=_s(p)
  if not p: return "Unknown"
  pl=p.lower()
  if any(k in pl for k in ["low","budget","entry"]): return "Low"
  if any(k in pl for k in ["mid","medium","standard"]): return "Mid"
  if any(k in pl for k in ["high","premium","enterprise"]): return "High"
  m=re.findall(r"\$?\s*([0-9][0-9,]*)", p)
  if m:
    v=int(m[0].replace(",",""))
    return "Low" if v<2000 else ("Mid" if v<10000 else "High")
  return p
def _warranty_months(w):
  w=_s(w).lower()
  if not w: return None
  m=re.search(r"(\d+(?:\.\d+)?)\s*(year|yr|years)", w)
  if m: return int(float(m.group(1))*12)
  m=re.search(r"(\d+(?:\.\d+)?)\s*(month|mo|months)", w)
  if m: return int(float(m.group(1)))
  return None
def _norm_compliance(c):
  c=_s(c)
  if not c: return "Unknown"
  parts=[_s(x) for x in re.split(r"[;,/|]+", c) if _s(x)]
  seen=set(); out=[]
  for p in parts:
    k=p.lower()
    if k not in seen:
      out.append(p); seen.add(k)
  return "; ".join(out) if out else "Unknown"
def _score_ops(rec):
  ops=[]
  comp=rec.get("compliance_posture","")
  if "Unknown" in comp or comp.strip()=="":
    ops.append("Lead with explicit compliance certifications + documentation.")
  elif not any(k in comp.upper() for k in ["UL","CE","FCC","ROHS","ISO"]):
    ops.append("Differentiate with recognized compliance marks (UL/CE/FCC/RoHS) and audit trail.")
  w=_warranty_months(rec.get("warranty_term"))
  if w is None:
    ops.append("Publish clear warranty terms and service SLAs.")
  elif w < 12:
    ops.append("Offer 12–24 month warranty and transparent RMA process.")
  if rec.get("pricing_band","") in {"High","Premium"}:
    ops.append("Compete with value bundle (training, spares, onboarding) and faster ROI proof.")
  if rec.get("supplier_type")=="Kit Supplier":
    ops.append("Add integration support, wiring diagrams, and installer network to reduce buyer risk.")
  if rec.get("supplier_type")=="Integrator":
    ops.append("Offer standardized kit SKUs + fixed-scope packages to reduce sales cycle.")
  if rec.get("supplier_type")=="OEM Program":
    ops.append("Create white-label pathway and partner enablement (co-marketing, tiered support).")
  if not ops: ops.append("Clarify ICP, quantify outcomes, and publish reference deployments.")
  return " ".join(dict.fromkeys(ops))
def _load_seed(path: Path):
  if path.exists():
    with path.open("r", encoding="utf-8-sig", newline="") as f:
      r=csv.DictReader(f)
      return [dict(row) for row in r]
  return [dict(x) for x in SEED_FALLBACK]
def _load_overrides(path: Path|None):
  if not path or not path.exists(): return {}
  txt=path.read_text(encoding="utf-8")
  try:
    import yaml
    data=yaml.safe_load(txt) or {}
    return data if isinstance(data, dict) else {}
  except Exception:
    # allow JSON as a fallback
    try:
      import json
      data=json.loads(txt)
      return data if isinstance(data, dict) else {}
    except Exception:
      return {}
def _apply_overrides(recs, ov):
  if not ov: return recs
  by_name={_s(r.get("competitor")): r for r in recs if _s(r.get("competitor"))}
  edits=ov.get("edit") or ov.get("updates") or {}
  adds=ov.get("add") or ov.get("new") or []
  deletes=set(_s(x) for x in (ov.get("delete") or []) if _s(x))
  for name, patch in (edits.items() if isinstance(edits, dict) else []):
    if name in by_name and isinstance(patch, dict):
      by_name[name].update(patch)
  for a in adds if isinstance(adds, list) else []:
    if isinstance(a, dict) and _s(a.get("competitor")):
      by_name.setdefault(_s(a["competitor"]), {}).update(a)
  for d in deletes:
    by_name.pop(d, None)
  return list(by_name.values())
def _normalize(recs):
  out=[]
  for r in recs:
    rec={k:_s(v) for k,v in (r or {}).items()}
    rec["competitor"]=rec.get("competitor") or rec.get("name") or "Unknown"
    rec["supplier_type"]=_norm_type(rec.get("supplier_type") or rec.get("type"))
    rec["region"]=rec.get("region") or rec.get("geo") or "Unknown"
    rec["pricing_band"]=_norm_pricing(rec.get("pricing_band") or rec.get("pricing") or rec.get("price_band"))
    rec["price_range_usd"]=rec.get("price_range_usd") or rec.get("price_range") or rec.get("pricing_notes") or ""
    rec["warranty_term"]=rec.get("warranty_term") or rec.get("warranty") or ""
    rec["compliance_posture"]=_norm_compliance(rec.get("compliance_posture") or rec.get("compliance") or "")
    rec["differentiators"]=rec.get("differentiators") or rec.get("differentiation") or rec.get("strengths") or ""
    rec["notes"]=rec.get("notes") or ""
    rec["opportunities"]=_score_ops(rec)
    out.append(rec)
  out.sort(key=lambda x:(x.get("supplier_type",""), x.get("competitor","")))
  return out
def _write_csv(path: Path, rows, cols):
  path.parent.mkdir(parents=True, exist_ok=True)
  with path.open("w", encoding="utf-8", newline="") as f:
    w=csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    for r in rows: w.writerow({c:r.get(c,"") for c in cols})
def _write_xlsx_or_csv(xlsx_path: Path, csv_path: Path, rows, cols):
  try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="Competitive Map"
    ws.append(cols)
    for r in rows: ws.append([r.get(c,"") for c in cols])
    for i,c in enumerate(cols, start=1):
      maxlen=max([len(str(c))]+[len(str(r.get(c,"") or "")) for r in rows]) if rows else len(str(c))
      ws.column_dimensions[get_column_letter(i)].width=min(60, max(10, maxlen+2))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))
    return ("xlsx", xlsx_path)
  except Exception:
    _write_csv(csv_path, rows, cols)
    return ("csv", csv_path)
def _summary_md(path: Path, rows, written_kind, written_path: Path):
  types={}
  bands={"Low":0,"Mid":0,"High":0,"Unknown":0}
  compli={"Known":0,"Unknown":0}
  warr={"Known":0,"Unknown":0}
  for r in rows:
    types[r["supplier_type"]]=types.get(r["supplier_type"],0)+1
    b=r.get("pricing_band","Unknown")
    bands[b]=bands.get(b,0)+1
    cm=r.get("compliance_posture","Unknown")
    compli["Unknown" if cm=="Unknown" else "Known"]+=1
    wm=_warranty_months(r.get("warranty_term",""))
    warr["Unknown" if wm is None else "Known"]+=1
  top_ops=[]
  for r in rows:
    for sent in re.split(r"(?<=\.)\s+", r.get("opportunities","").strip()):
      s=_s(sent)
      if s: top_ops.append(s)
  # de-dupe preserve order
  seen=set(); top=[]
  for s in top_ops:
    k=s.lower()
    if k not in seen:
      seen.add(k); top.append(s)
    if len(top)>=8: break
  lines=[]
  lines.append(f"# Competitive landscape summary\n")
  lines.append(f"- Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
  lines.append(f"- Records: {len(rows)}")
  lines.append(f"- Matrix: `{written_path.name}` ({written_kind.upper()})\n")
  lines.append("## Coverage by supplier type")
  for k in sorted(types, key=lambda x:(x!="Kit Supplier", x!="Integrator", x!="OEM Program", x)):
    lines.append(f"- {k}: {types[k]}")
  lines.append("\n## Pricing bands")
  for k in ["Low","Mid","High","Unknown"]:
    if k in bands: lines.append(f"- {k}: {bands.get(k,0)}")
  lines.append("\n## Warranty / compliance disclosure")
  lines.append(f"- Warranty disclosed: {warr['Known']} / {len(rows)}")
  lines.append(f"- Compliance disclosed: {compli['Known']} / {len(rows)}\n")
  lines.append("## Differentiation opportunities (aggregate)")
  for s in top: lines.append(f"- {s}")
  lines.append("\n## Notes")
  lines.append("- Treat this as a working matrix; update using the overrides file to refine pricing, warranty, and compliance specifics.")
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_text("\n".join(lines).strip()+"\n", encoding="utf-8")

def main():
  ap=argparse.ArgumentParser(description="Generate competitive landscape matrix and summary.")
  ap.add_argument("--seed", type=str, default=str(DEFAULT_SEED), help="Path to competitors_seed.csv")
  ap.add_argument("--overrides", type=str, default="", help="Path to overrides YAML/JSON (optional)")
  ap.add_argument("--out-dir", type=str, default=str(DEFAULT_OUT_DIR), help="Output directory")
  ap.add_argument("--xlsx", type=str, default="", help="Output XLSX path (optional)")
  ap.add_argument("--csv", type=str, default="", help="Output CSV fallback path (optional)")
  ap.add_argument("--md", type=str, default="", help="Output Markdown summary path (optional)")
  args=ap.parse_args()

  out_dir=Path(args.out_dir)
  seed=Path(args.seed)
  ov=Path(args.overrides) if args.overrides else None
  xlsx=Path(args.xlsx) if args.xlsx else (out_dir/"competitive_map.xlsx")
  csvp=Path(args.csv) if args.csv else (out_dir/"competitive_map.csv")
  md=Path(args.md) if args.md else (out_dir/"competitive_summary.md")

  recs=_load_seed(seed)
  recs=_apply_overrides(recs, _load_overrides(ov))
  rows=_normalize(recs)

  cols=["competitor","supplier_type","region","pricing_band","price_range_usd","warranty_term","compliance_posture","differentiators","opportunities","notes"]
  kind, path=_write_xlsx_or_csv(xlsx, csvp, rows, cols)
  _summary_md(md, rows, kind, path)

if __name__=="__main__":
  main()
